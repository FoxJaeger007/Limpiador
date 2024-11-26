import pandas as pd
import logging
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

def apply_styles(ws, min_row, max_row, min_col, max_col, bold=False):
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if bold:
                cell.font = Font(bold=True)
            cell.border = border

def auto_adjust_column_width(ws, exclude_first_col=True):
    for col in ws.columns:
        if exclude_first_col and col[0].column_letter == 'A':
            continue
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 5
        ws.column_dimensions[column].width = adjusted_width

def log_null_data(df, ws, excel_filename):
    if df.empty:
        ws.append(["El DataFrame está vacío."])
        return

    ws.append(["***** Procesando archivo: {} *****".format(excel_filename)])

    # Registro de datos nulos
    null_data = df.isnull().sum()
    ws.append([])
    ws.append(["Datos nulos por columna:"])

    # Aplicar negrilla a los títulos de las tablas inmediatamente y agregar bordes
    apply_styles(ws, ws.max_row, ws.max_row, 1, 2, bold=True)

    # Detalle de datos nulos con bordes
    for column, count in null_data.items():
        ws.append([column, count])
        apply_styles(ws, ws.max_row, ws.max_row, 1, 2)

    # Ajustar el ancho de la columna A según el contenido de los datos nulos, excluyendo la primera fila
    max_length = 0
    for cell in ws['A'][1:]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions['A'].width = adjusted_width

    ws.append([])
    ws.append([])

    # Generar listado de filas con datos nulos
    null_rows = df[df.isnull().any(axis=1)]
    if not null_rows.empty:
        ws.append(["Detalle de filas con datos nulos:"])

        # Aplicar negrilla al título "Detalle de filas con datos nulos"
        apply_styles(ws, ws.max_row, ws.max_row, 1, 2, bold=True)

        headers = list(null_rows.columns)
        ws.append(headers)
        apply_styles(ws, ws.max_row, ws.max_row, 1, len(headers), bold=True)
        for row in null_rows.itertuples(index=False):
            ws.append(list(row))
            apply_styles(ws, ws.max_row, ws.max_row, 1, len(headers))

    ws.append([])

    # Autoajustar el ancho de las columnas, exceptuando la primera columna
    auto_adjust_column_width(ws)

def log_duplicate_data(df, ws, column_prefix):
    if df.empty:
        ws.append(["El DataFrame está vacío."])
        return

    # Encontrar columnas con el prefijo especificado
    columns_with_prefix = [col for col in df.columns if col.startswith(column_prefix)]
    ws.append(["Columnas con el prefijo '{}'".format(column_prefix)])
    apply_styles(ws, ws.max_row, ws.max_row, 1, 1, bold=True)

    ws.append(columns_with_prefix)
    apply_styles(ws, ws.max_row, ws.max_row, 1, len(columns_with_prefix))

    for column_name in columns_with_prefix:
        df[column_name] = df[column_name].astype(str).str.lower()
        ws.append([])

        # Registro de datos duplicados en cada columna con el prefijo
        duplicate_data = df.duplicated(subset=[column_name]).sum()
        ws.append(["Número de filas duplicadas en la columna '{}': {}".format(column_name, duplicate_data)])
        apply_styles(ws, ws.max_row, ws.max_row, 1, 1, bold=True)

        if duplicate_data > 0:
            ws.append(["Filas duplicadas en la columna '{}':".format(column_name)])
            apply_styles(ws, ws.max_row, ws.max_row, 1, 1, bold=True)

            duplicated_rows = df[df.duplicated(subset=[column_name], keep=False)]

            headers = list(duplicated_rows.columns)
            ws.append(headers)
            apply_styles(ws, ws.max_row, ws.max_row, 1, len(headers), bold=True)

            for row in duplicated_rows.itertuples(index=False):
                ws.append(list(row))
                apply_styles(ws, ws.max_row, ws.max_row, 1, len(headers))

    ws.append([])

    # Autoajustar el ancho de las columnas, exceptuando la primera columna
    auto_adjust_column_width(ws)

def process_excel_files_in_folder(folder_path, column_prefix, output_excel_filename):
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xls')]

    if not excel_files:
        print("No se encontraron archivos Excel en la carpeta especificada.")
        return

    wb = Workbook()
    wb.remove(wb.active)  # Eliminar la hoja de trabajo por defecto

    for excel_file in excel_files:
        excel_path = os.path.join(folder_path, excel_file)
        sheet_name = os.path.splitext(excel_file)[0][:31]

        try:
            df = pd.read_excel(excel_path)
            print(f"Archivo '{excel_path}' leído exitosamente.")

            ws = wb.create_sheet(title=sheet_name)

            log_null_data(df, ws, excel_file)
            log_duplicate_data(df, ws, column_prefix)

            print(f"Registro completado para '{excel_file}'. Revisa el archivo '{output_excel_filename}' para los detalles.")
        except FileNotFoundError:
            print(f"Error: El archivo '{excel_path}' no fue encontrado.")
        except Exception as e:
            print(f"Ocurrió un error al procesar el archivo '{excel_path}': {e}")

    wb.save(output_excel_filename)

def main():
    if len(sys.argv) < 4:
        print("Uso: python programa.py carpeta_de_excel archivo_salida prefijo")
        sys.exit(1)

    folder_path = sys.argv[1]
    output_excel_filename = sys.argv[2]
    column_prefix = sys.argv[3]

    if not os.path.isdir(folder_path):
        print(f"Error: La carpeta '{folder_path}' no existe o no es un directorio.")
        sys.exit(1)

    process_excel_files_in_folder(folder_path, column_prefix, output_excel_filename)

if __name__ == "__main__":
    main()
